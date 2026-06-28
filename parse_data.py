"""
Parse ttt.xlsx into a normalized JSON data model.

Output model:
{
  "grades": [
    {
      "grade": "Grade 6",
      "classes": ["6A",...],
      "buckets": [ { "id": "...", "color": "...", "subjects": [...] } ],   # same-period groups
      "subjects": [
         {
           "subject": "English", "weekly": 7, "color": "...", "bucketId": <id|null>,
           "religionBlock": false,
           "assignments": [ {"teacher":"Ms. Nimnada","classes":["6A","6B","6E","6F"],"periods":7}, ... ],
           "mergedGroups": [["6A","6B"],...]  # for bucket merged classes (optional)
         }
      ]
    }, ...
  ]
}
The generator (in JS) consumes this. Colour == bucket key. "religionBlock" subjects all share one bucket.
"""
import json, re
from openpyxl import load_workbook

SRC = "ttt.xlsx"

def color_key(cell):
    f = cell.fill
    if not f or not f.patternType:
        return "none"
    c = f.start_color
    if c.type == "rgb":
        return "rgb:" + str(c.rgb)
    if c.type == "theme":
        return "t%s_%.2f" % (c.theme, c.tint or 0)
    if c.type == "indexed":
        return "idx:" + str(c.indexed)
    return "x"

def clean(s):
    if s is None:
        return None
    return re.sub(r"\s+", " ", str(s)).strip()

def subj_name_and_weekly(raw):
    # "English (7)" -> ("English", 7)
    raw = clean(raw)
    m = re.search(r"\((\d+)\)\s*$", raw)
    weekly = int(m.group(1)) if m else None
    name = re.sub(r"\s*\(\d+\)\s*$", "", raw).strip()
    return name, weekly

RELIGIONS = {"Buddhism", "Hinduism", "Islam", "Catholicism", "Christianity"}

# Teacher overrides: (grade, original_teacher) -> new_teacher.
# Rule: Grade 6 "New MAT" Maths periods are taught by Mr. Mithun (who also teaches G8 Maths).
TEACHER_OVERRIDES = {
    ("Grade 6", "New MAT"): "Mr. Mithun",
}

# Period-count corrections: (grade, subject) -> weekly periods. Lets the school
# balance a grade to fit the week. Grade 8 Computer Science 3->2 so that, with
# Religion (1) included and assembly as fixed Period 1, the week totals 39.
PERIOD_OVERRIDES = {
    ("Grade 8", "Computer Science"): 2,
}

# Period-count overrides: (grade, subject) -> weekly periods. Applied during parsing.
# Grade 8 Computer Science 3->2 so the grade fits 39 teaching slots (assembly = P1).
PERIOD_OVERRIDES = {
    ("Grade 8", "Computer Science"): 2,
}

def localize_placeholder(teacher, grade_name):
    """Placeholder teachers ('New ...') are unfilled positions; give each grade its own
    so they never clash across grades (e.g. 'New ENG' -> 'New ENG (Grade 9)'). Real
    named teachers and the Mr. Mithun override are left unchanged."""
    if teacher and teacher.strip().lower().startswith("new"):
        return f"{teacher} ({grade_name})"
    return teacher

def get_row_merge_groups(ws, row, hdr_cols):
    """For a given subject row, return the list of merge-groups (tuples of class names)
    that are explicitly merged together in the sheet. Classes not part of any merge
    appear as singletons. This captures the TRUE combined-class structure, e.g.
    Grade 6 Cultural Dancing -> [(6A,6B),(6C,6D),(6E,6F,6G)]."""
    merged_cols = {}  # col -> group tuple
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= row <= mr.max_row and mr.min_col >= 3 and mr.max_col > mr.min_col:
            cols = [c for c in range(mr.min_col, mr.max_col + 1) if c in hdr_cols]
            grp = tuple(hdr_cols[c] for c in cols)
            if len(grp) > 1:
                for c in cols:
                    merged_cols[c] = grp
    return merged_cols

def parse_simple_grade(ws, grade_name, class_cols, class_row=4):
    classes = []
    col_index = {}
    for col in class_cols:
        v = clean(ws.cell(class_row, col).value)
        if v:
            classes.append(v)
            col_index[col] = v
    subjects = {}
    for r in range(class_row + 1, ws.max_row + 1):
        raw = ws.cell(r, 1).value
        if raw is None:
            continue
        name, weekly = subj_name_and_weekly(raw)
        if not name:
            continue
        if (grade_name, name) in PERIOD_OVERRIDES:
            weekly = PERIOD_OVERRIDES[(grade_name, name)]
        teacher = clean(ws.cell(r, 2).value)
        teacher = TEACHER_OVERRIDES.get((grade_name, teacher), teacher)
        teacher = localize_placeholder(teacher, grade_name)
        color = color_key(ws.cell(r, 1))
        povr = PERIOD_OVERRIDES.get((grade_name, name))

        # Determine this row's merge-groups so combined classes are taught together.
        row_merges = get_row_merge_groups(ws, r, col_index)

        # Read the period value for each class column. For a merged range, openpyxl
        # stores the value only in the top-left cell, so propagate it to the group.
        cls_periods = {}
        for col, cls in col_index.items():
            val = ws.cell(r, col).value
            if isinstance(val, (int, float)) and val:
                cls_periods[cls] = povr if povr is not None else int(val)
        # propagate merged value across each merge-group
        for col, grp in row_merges.items():
            # find any value present in the group
            gval = None
            for c in range(min(k for k, v in col_index.items() if v in grp),
                           max(k for k, v in col_index.items() if v in grp) + 1):
                v = ws.cell(r, c).value
                if isinstance(v, (int, float)) and v:
                    gval = povr if povr is not None else int(v)
                    break
            if gval is not None:
                for cls in grp:
                    cls_periods[cls] = gval

        if povr is not None and weekly is not None:
            weekly = povr
        is_religion = name in RELIGIONS
        key = name
        if key not in subjects:
            subjects[key] = {
                "subject": name, "weekly": weekly, "color": color,
                "religionBlock": is_religion, "assignments": [],
            }
        if subjects[key]["weekly"] is None and weekly:
            subjects[key]["weekly"] = weekly

        if cls_periods:
            # Build assignments by merge-group: each combined group becomes ONE
            # assignment (taught together, one slot, one teacher). Unmerged classes
            # are singletons.
            handled = set()
            groups = []
            seen_groups = set()
            for cls in cls_periods:
                col = next(k for k, v in col_index.items() if v == cls)
                if col in row_merges:
                    grp = tuple(c for c in row_merges[col] if c in cls_periods)
                    if grp and grp not in seen_groups:
                        seen_groups.add(grp)
                        groups.append(list(grp))
                        handled.update(grp)
                elif cls not in handled:
                    groups.append([cls])
                    handled.add(cls)
            for grp in groups:
                per = max(cls_periods[c] for c in grp)
                subjects[key]["assignments"].append({
                    "teacher": teacher, "classes": grp, "periods": per,
                    "mergedGroup": grp if len(grp) > 1 else None,
                })
        else:
            subjects[key]["assignments"].append({
                "teacher": teacher, "classes": [], "periods": weekly or 1,
            })
    return classes, list(subjects.values())

def parse_grade_9_10(ws):
    """Grade 9-10 sheet: columns are streams. Grade 9 = cols 3..8, Grade 10 = cols 9..14."""
    g9_cols, g10_cols = {}, {}
    seen = {}
    def uniq(name):
        if name in seen:
            seen[name] += 1
            # rename duplicate trailing letter (e.g. "9 NAT A" -> "9 NAT B")
            if name and name[-1] in "AB":
                return name[:-1] + chr(ord(name[-1]) + seen[name])
            return f"{name} ({seen[name]})"
        seen[name] = 0
        return name
    for col in range(3, 9):
        v = clean(ws.cell(4, col).value)
        if v: g9_cols[col] = uniq(v)
    for col in range(9, 15):
        v = clean(ws.cell(4, col).value)
        if v: g10_cols[col] = uniq(v)

    def collect(cols, grade_name):
        subjects = {}
        for r in range(5, ws.max_row + 1):
            raw = ws.cell(r, 1).value
            if raw is None:
                continue
            name, weekly = subj_name_and_weekly(raw)
            if not name:
                continue
            teacher = clean(ws.cell(r, 2).value)
            teacher = re.sub(r"\s*\(.*?\)\s*", "", teacher).strip() if teacher else teacher
            teacher = localize_placeholder(teacher, grade_name)
            teacher = TEACHER_OVERRIDES.get((grade_name, teacher), teacher)
            color = color_key(ws.cell(r, 1))
            cls_periods = {}
            for col, cls in cols.items():
                val = ws.cell(r, col).value
                if isinstance(val, (int, float)) and val:
                    cls_periods[cls] = int(val)
            if not cls_periods:
                continue
            is_religion = False  # 9-10 religion is colour-bucketed per stream (t6=PED@1, t3=NAT@2)
            key = name + "|" + (teacher or "") + "|" + color
            entry = subjects.get(key)
            if not entry:
                entry = subjects[key] = {
                    "subject": name, "weekly": weekly, "color": color,
                    "religionBlock": is_religion, "assignments": [],
                }
            # group classes that share the SAME period count into separate assignments
            by_count = {}
            for c, p in cls_periods.items():
                by_count.setdefault(p, []).append(c)
            for p, cls_list in by_count.items():
                entry["assignments"].append({
                    "teacher": teacher, "classes": cls_list, "periods": p,
                })
        merged = {}
        for s in subjects.values():
            mkey = s["subject"] + "@" + s["color"]
            m = merged.get(mkey)
            if not m:
                merged[mkey] = s
            else:
                m["assignments"].extend(s["assignments"])
        return list(cols.values()), list(merged.values())

    out = []
    for cols, gname in [(g9_cols, "Grade 9"), (g10_cols, "Grade 10")]:
        classes, subjects = collect(cols, gname)
        # De-duplicate: if a subject is BOTH a core (white) subject and a coloured
        # bucket subject for the SAME stream, the bucket entry is a colour-bleed
        # duplicate (e.g. NAT English Literature appears as core@3 and t8@3). Remove
        # that stream from the bucket assignment so it is counted once (as core).
        core_by_class = {}
        for s in subjects:
            if s["color"] in ("white", "none", "idx:9"):
                for a in s["assignments"]:
                    for c in a["classes"]:
                        core_by_class.setdefault(c, set()).add(s["subject"])
        for s in subjects:
            if s["color"] in ("white", "none", "idx:9"):
                continue
            for a in s["assignments"]:
                a["classes"] = [c for c in a["classes"]
                                if s["subject"] not in core_by_class.get(c, set())]
            s["assignments"] = [a for a in s["assignments"] if a["classes"]]
        subjects = [s for s in subjects if s["assignments"]]
        out.append((gname, classes, subjects))
    return out

def assign_buckets(subjects):
    NON_BUCKET = {"none", "idx:9", "i9", "rgb:FFFFFFFF", "rgb:00000000"}
    buckets = {}
    for s in subjects:
        col = s["color"]
        if s["religionBlock"]:
            col = "RELIGION"
        if col not in NON_BUCKET:
            buckets.setdefault(col, []).append(s["subject"])
            s["bucketId"] = col
        else:
            s["bucketId"] = None
    return [{"id": k, "color": k, "subjects": v} for k, v in buckets.items()]

def parse_grade_11_12(ws):
    """Grade 11-12 A/L sheet. Columns 4-7 = Grade 11 streams, 8-11 = Grade 12 streams.
    Column B = weekly period count. Subjects in rows 5-22 (A/L mains); 25-33 NAT
    activities; 36-45 PED activities. Data is taken as-is for later editing.
    Activity rows whose name lists several options (EM/WM/CUD/Art) are buckets."""
    g11 = {4: "11 PED SCI", 5: "11 PED COM", 6: "11 NAT SCI", 7: "11 NAT COM"}
    g12 = {8: "12 PED SCI", 9: "12 NAT SCI", 10: "12 PED COM", 11: "12 NAT COM"}

    def period_of(r, col):
        b = ws.cell(r, 2).value
        v = ws.cell(r, col).value
        try:
            return int(b)
        except (TypeError, ValueError):
            try:
                return int(v)
            except (TypeError, ValueError):
                return 1

    BUCKET_NAMES = {"EM , WM, CUD, Art", "Religion", "Religion "}

    def collect(cols, gname):
        subjects = {}
        for r in range(5, 46):
            raw = ws.cell(r, 1).value
            if not raw or str(raw).strip() in ("National", "PED", "Subject"):
                continue
            name = clean(raw)
            if name == "Assembly":
                continue  # assembly handled separately as fixed P1
            teacher = clean(ws.cell(r, 3).value)
            teacher = re.sub(r"\s*\(.*?\)\s*", "", teacher).strip() if teacher else teacher
            teacher = localize_placeholder(teacher, gname)
            cls_periods = {}
            for col, cls in cols.items():
                v = ws.cell(r, col).value
                if v is not None and str(v).strip():
                    cls_periods[cls] = period_of(r, col)
            if not cls_periods:
                continue
            is_bucket = any(name.startswith(b) for b in ("EM ", "Religion"))
            color = "ALBUCKET" if is_bucket else "none"
            key = name + "|" + (teacher or "") + "|" + color
            entry = subjects.get(key)
            if not entry:
                entry = subjects[key] = {
                    "subject": name, "weekly": None, "color": color,
                    "religionBlock": name.startswith("Religion"), "assignments": [],
                }
            by_count = {}
            for c, p in cls_periods.items():
                by_count.setdefault(p, []).append(c)
            for p, cl in by_count.items():
                entry["assignments"].append({"teacher": teacher, "classes": cl, "periods": p})
                if entry["weekly"] is None:
                    entry["weekly"] = p
        return list(cols.values()), list(subjects.values())

    out = []
    for cols, gname in [(g11, "Grade 11"), (g12, "Grade 12")]:
        classes, subjects = collect(cols, gname)
        out.append((gname, classes, subjects))
    return out

def build():
    wb = load_workbook(SRC)
    model = {"grades": []}

    simple = {
        "Grade 6": list(range(3, 11)),
        "Grade 7": list(range(3, 9)),
        "Grade 8": list(range(3, 9)),
    }
    for g, cols in simple.items():
        ws = wb[g]
        classes, subjects = parse_simple_grade(ws, g, cols)
        buckets = assign_buckets(subjects)
        model["grades"].append({
            "grade": g, "classes": classes, "subjects": subjects, "buckets": buckets,
        })

    # Grades 9 and 10 from the combined stream sheet
    for gname, classes, subjects in parse_grade_9_10(wb["Grade 9-10"]):
        buckets = assign_buckets(subjects)
        model["grades"].append({
            "grade": gname, "classes": classes, "subjects": subjects, "buckets": buckets,
        })

    # Grades 11 and 12 (A/L). Data taken as-is for editing; may exceed weekly capacity.
    for gname, classes, subjects in parse_grade_11_12(wb["Grade 11,12"]):
        buckets = assign_buckets(subjects)
        model["grades"].append({
            "grade": gname, "classes": classes, "subjects": subjects, "buckets": buckets,
        })
    with open("/home/claude/tt-app/data_model.json", "w") as f:
        json.dump(model, f, indent=2, ensure_ascii=False)
    # summary
    for g in model["grades"]:
        print(g["grade"], "classes", g["classes"])
        print("  buckets:")
        for b in g["buckets"]:
            print("   ", b["id"], b["subjects"])
    print("WROTE data_model.json")

if __name__ == "__main__":
    build()
